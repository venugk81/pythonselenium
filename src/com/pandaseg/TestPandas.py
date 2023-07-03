import openpyxl as pyxl

fil = "C:\\Users\\gopve\\Desktop\\venu\\practice py\\data_file.xlsx"

wb = pyxl.load_workbook(fil)
lst = wb.sheetnames
print(lst)

sheet = wb['Sheet1']

# https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/
print(sheet.title)
print(sheet["A1"].value)
for value in sheet.iter_rows(
        min_row=1, max_row=3, min_col=1, max_col=3,
        values_only=True):
    print(value)
# https://www.blog.pythonlibrary.org/2021/07/20/reading-spreadsheets-with-openpyxl-and-python/
